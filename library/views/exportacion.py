import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib.auth import get_user_model
from django.http import HttpResponse

from ..models import Libro
from ..permissions import admin_required

User = get_user_model()


# ─────────────────────────────────────────────────────────────────────────────
# F3 — Exportar Listados como Excel
# ─────────────────────────────────────────────────────────────────────────────

@admin_required
def exportar_usuarios_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    encabezados = ['Nombre', 'Email', 'Rol', 'Categorías', 'Activo']
    ws.append(encabezados)
    hf  = Font(bold=True, color="FFFFFF")
    hfi = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    ha  = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hfi; cell.alignment = ha

    usuarios = User.objects.select_related('rol').prefetch_related('categoriapermitida_set__categoria')
    for u in usuarios:
        cats = ', '.join(cp.categoria.nombre for cp in u.categoriapermitida_set.all())
        ws.append([
            u.name or '',
            u.email,
            u.rol.nombre if u.rol_id else '',
            cats,
            'Sí' if u.is_active else 'No',
        ])

    for i, w in enumerate([25, 35, 10, 40, 8], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="usuarios.xlsx"'
    return response


@admin_required
def exportar_libros_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libros"

    encabezados = ['Título', 'Autor', 'Año', 'Categoría', 'ISBN', 'Descripción']
    ws.append(encabezados)
    hf  = Font(bold=True, color="FFFFFF")
    hfi = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    ha  = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font = hf; cell.fill = hfi; cell.alignment = ha

    libros = Libro.objects.select_related('categoria').order_by('titulo')
    for libro in libros:
        ws.append([
            libro.titulo,
            libro.autor,
            libro.año,
            libro.categoria.nombre if libro.categoria_id else '',
            libro.isbn or '',
            libro.descripcion or '',
        ])

    for i, w in enumerate([35, 30, 8, 20, 18, 50], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="libros.xlsx"'
    return response

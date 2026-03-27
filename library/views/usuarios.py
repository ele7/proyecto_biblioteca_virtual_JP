import csv
import io
import logging
import secrets

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from ..forms import CustomUserChangeForm, UploadUsersForm, UsuarioForm
from ..models import Categoria, CategoriaPermitida, CustomUser, Rol
from ..permissions import admin_required
from ..services.email_service import enviar_bienvenida, enviar_notificacion_password
from .auditoria import _registrar_auditoria

logger = logging.getLogger(__name__)
User = get_user_model()


# ─────────────────────────────────────────────────────────────────────────────
# Usuarios — solo ADMIN
# ─────────────────────────────────────────────────────────────────────────────

@admin_required
def usuarios_listar(request):
    usuarios = User.objects.select_related('rol').prefetch_related(
        'categoriapermitida_set__categoria'
    )
    usuarios_con_categorias = [
        {
            'usuario':    u,
            'categorias': [cp.categoria.nombre for cp in u.categoriapermitida_set.all()],
        }
        for u in usuarios
    ]
    return render(request, "maestros/usuarios/listar.html", {
        "usuarios_con_categorias": usuarios_con_categorias,
    })


@admin_required
def usuarios_crear(request):
    categorias = Categoria.objects.all()

    if request.method == "POST":
        form = UsuarioForm(request.POST)
        categorias_seleccionadas = request.POST.getlist('categorias')

        if form.is_valid():
            password_plain = form.cleaned_data['password']
            user = form.save(commit=False)
            user.set_password(password_plain)
            user.save()

            cats = Categoria.objects.filter(id__in=categorias_seleccionadas)
            CategoriaPermitida.objects.bulk_create(
                [CategoriaPermitida(usuario=user, categoria=cat) for cat in cats],
                ignore_conflicts=True,
            )

            _registrar_auditoria(request, 'CREAR', 'CustomUser', str(user))
            enviar_bienvenida(user, password_plain)
            messages.success(request, "Usuario creado correctamente.")
            return redirect("library:usuarios_listar")
    else:
        form = UsuarioForm()

    return render(request, "maestros/usuarios/crear.html", {
        "form":               form,
        "categorias":         categorias,
        "categorias_usuario": [],
    })


@admin_required
def usuarios_editar(request, pk):
    usuario = get_object_or_404(CustomUser, pk=pk)

    if request.method == "POST":
        form = CustomUserChangeForm(request.POST, instance=usuario)
        if form.is_valid():
            user = form.save(commit=False)
            pwd = form.cleaned_data.get('password')
            if pwd:
                user.set_password(pwd)
            user.save()

            categorias_seleccionadas = form.cleaned_data['categorias']
            CategoriaPermitida.objects.filter(usuario=user).exclude(
                categoria__in=categorias_seleccionadas
            ).delete()
            for cat in categorias_seleccionadas:
                CategoriaPermitida.objects.get_or_create(usuario=user, categoria=cat)

            _registrar_auditoria(request, 'EDITAR', 'CustomUser', str(user))
            if pwd:
                enviar_notificacion_password(user, pwd)
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect("library:usuarios_listar")
    else:
        categorias_actuales = CategoriaPermitida.objects.filter(
            usuario=usuario
        ).values_list('categoria', flat=True)
        form = CustomUserChangeForm(instance=usuario)
        form.fields['categorias'].initial = categorias_actuales

    return render(request, "maestros/usuarios/editar.html", {
        "form":    form,
        "usuario": usuario,
    })


@admin_required
def usuarios_eliminar(request, pk):
    usuario = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        if usuario == request.user:
            messages.error(request, 'No puedes eliminar tu propio usuario.')
            return redirect('library:usuarios_listar')
        nombre = str(usuario)
        usuario.delete()
        _registrar_auditoria(request, 'ELIMINAR', 'CustomUser', nombre)
        messages.success(request, f'Usuario "{nombre}" eliminado correctamente.')
    return redirect('library:usuarios_listar')


@admin_required
def carga_masiva_usuarios(request):
    """
    GET:  Muestra el formulario vacío con instrucciones y botón de plantilla.
    POST: Procesa el archivo y muestra el resumen detallado en la misma página.
    """
    resultado = None
    form = UploadUsersForm()

    if request.method == "POST":
        form = UploadUsersForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES["archivo"]
            ext     = archivo.name.split('.')[-1].lower()
            data_rows = []

            resultado = {
                'creados':      0,
                'duplicados':   [],
                'errores':      [],
                'advertencias': [],
            }

            if ext == "csv":
                try:
                    decoded_file = archivo.read().decode("utf-8").splitlines()
                    data_rows    = list(csv.DictReader(decoded_file))
                except Exception as exc:
                    resultado['errores'].append(f"No se pudo leer el CSV: {exc}")
            elif ext in ["xlsx", "xls"]:
                try:
                    df        = pd.read_excel(archivo, dtype=str)
                    data_rows = df.to_dict(orient="records")
                except Exception as exc:
                    resultado['errores'].append(f"No se pudo leer el Excel: {exc}")
            else:
                resultado['errores'].append(
                    "Formato no soportado. Usa CSV (.csv) o Excel (.xlsx)."
                )

            for fila_num, row in enumerate(data_rows, start=2):
                email      = str(row.get("email", "")).strip()
                name       = str(row.get("name", "")).strip()
                rol_nombre = str(row.get("rol", "")).strip()
                password   = str(row.get("password", "")).strip()

                if not email:
                    resultado['errores'].append(
                        f"Fila {fila_num}: email vacío. Fila omitida."
                    )
                    continue

                if not rol_nombre:
                    resultado['errores'].append(
                        f"Fila {fila_num} ('{email}'): falta la columna 'rol'. Fila omitida."
                    )
                    continue

                if User.objects.filter(email=email).exists():
                    resultado['duplicados'].append(
                        f"Fila {fila_num}: '{email}' ya existe en el sistema."
                    )
                    continue

                try:
                    rol_obj = Rol.objects.get(nombre=rol_nombre)
                except Rol.DoesNotExist:
                    resultado['errores'].append(
                        f"Fila {fila_num} ('{email}'): el rol '{rol_nombre}' no existe. "
                        f"Valores válidos: ADMIN, LECTOR."
                    )
                    continue

                if not password:
                    password = secrets.token_urlsafe(12)
                    resultado['advertencias'].append(
                        f"Fila {fila_num} ('{email}'): sin contraseña, "
                        f"se generó una temporal automáticamente."
                    )

                try:
                    with transaction.atomic():
                        user = User.objects.create_user(
                            name=name,
                            email=email,
                            password=password,
                            rol=rol_obj,
                        )

                        nombres_categorias = [
                            c.strip()
                            for c in str(row.get("categorias", "")).split(",")
                            if c.strip()
                        ]
                        if nombres_categorias:
                            cats = Categoria.objects.filter(nombre__in=nombres_categorias)
                            encontradas = set(cats.values_list('nombre', flat=True))
                            no_encontradas = set(nombres_categorias) - encontradas
                            if no_encontradas:
                                resultado['advertencias'].append(
                                    f"Fila {fila_num} ('{email}'): categoría(s) no encontrada(s): "
                                    f"{', '.join(sorted(no_encontradas))}. Se omitieron."
                                )
                            CategoriaPermitida.objects.bulk_create(
                                [CategoriaPermitida(usuario=user, categoria=cat) for cat in cats],
                                ignore_conflicts=True,
                            )

                        resultado['creados'] += 1
                        enviar_bienvenida(user, password)

                except Exception as exc:
                    resultado['errores'].append(
                        f"Fila {fila_num} ('{email}'): error inesperado al guardar: {exc}"
                    )

    return render(request, "maestros/usuarios/carga_masiva.html", {
        "form":      form,
        "resultado": resultado,
    })


@admin_required
def descargar_plantilla_usuarios(request):
    """Genera y descarga en memoria una plantilla Excel para la carga masiva de usuarios."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Usuarios"

    encabezados = ['name', 'email', 'rol', 'password', 'categorias']
    ws.append(encabezados)

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.append([
        'Juan Pérez',
        'juan.perez@ejemplo.com',
        'LECTOR',
        'micontrasena123',
        'Teología, Historia',
    ])

    anchos = [25, 35, 10, 20, 40]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws2 = wb.create_sheet(title="Instrucciones")
    instrucciones = [
        ("Campo",       "¿Requerido?", "Descripción"),
        ("name",        "NO",  "Nombre completo del usuario (máx. 150 caracteres)."),
        ("email",       "SÍ",  "Correo electrónico del usuario. Debe ser único en el sistema."),
        ("rol",         "SÍ",  "Rol del usuario. Valores posibles: ADMIN o LECTOR."),
        ("password",    "NO",  "Contraseña inicial. Si se deja vacía se generará una automáticamente."),
        ("categorias",  "NO",  "Nombres de categorías separados por coma. Deben existir en el sistema."),
    ]
    for fila in instrucciones:
        ws2.append(fila)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 70

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_usuarios.xlsx"'
    return response

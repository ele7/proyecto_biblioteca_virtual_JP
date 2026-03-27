import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify

from ..forms import CargaMasivaLibrosForm, LibroForm
from ..models import Categoria, EstadisticaUsuario, HistorialLectura, Libro, ProgresoLectura
from ..permissions import admin_required, es_admin, usuario_puede_ver_categoria, usuario_puede_ver_libro
from ..services.email_service import enviar_notificacion_libro
from .auditoria import _registrar_auditoria

_ORDER_MAP = {'az': 'titulo', 'za': '-titulo', 'reciente': '-año', 'antiguo': 'año'}


# ─────────────────────────────────────────────────────────────────────────────
# Libros — solo ADMIN
# ─────────────────────────────────────────────────────────────────────────────

@admin_required
def listar_libros(request):
    libros       = Libro.objects.select_related('categoria')
    q            = request.GET.get('q', '').strip()
    categoria_id = request.GET.get('categoria', '').strip()
    year_desde   = request.GET.get('year_desde', '').strip()
    year_hasta   = request.GET.get('year_hasta', '').strip()
    orden        = request.GET.get('orden', 'az').strip()

    if q:
        libros = libros.filter(Q(titulo__icontains=q) | Q(autor__icontains=q))
    if categoria_id:
        libros = libros.filter(categoria_id=categoria_id)
    if year_desde:
        try:
            libros = libros.filter(año__gte=int(year_desde))
        except ValueError:
            pass
    if year_hasta:
        try:
            libros = libros.filter(año__lte=int(year_hasta))
        except ValueError:
            pass
    libros = libros.order_by(_ORDER_MAP.get(orden, 'titulo'))

    return render(request, 'maestros/libros/listar_libros.html', {
        'libros':                 libros,
        'categorias_disponibles': Categoria.objects.all(),
        'filtros': {
            'q':            q,
            'categoria_id': categoria_id,
            'year_desde':   year_desde,
            'year_hasta':   year_hasta,
            'orden':        orden,
        },
    })


@admin_required
def agregar_libro(request):
    if request.method == "POST":
        form = LibroForm(request.POST, request.FILES)
        if form.is_valid():
            libro_nuevo = form.save()
            _registrar_auditoria(request, 'CREAR', 'Libro', str(libro_nuevo))
            enviar_notificacion_libro(libro_nuevo)
            messages.success(request, f'Libro "{libro_nuevo.titulo}" agregado correctamente.')
            return redirect('library:listar_libros')
    else:
        form = LibroForm()
    return render(request, 'maestros/libros/agregar_libro.html', {'form': form})


@admin_required
def editar_libro(request, libro_id):
    libro = get_object_or_404(Libro, id=libro_id)

    if request.method == 'POST':
        form = LibroForm(request.POST, request.FILES, instance=libro)
        if form.is_valid():
            libro_edit = form.save()
            _registrar_auditoria(request, 'EDITAR', 'Libro', str(libro_edit))
            messages.success(request, 'Libro actualizado correctamente.')
            return redirect('library:listar_libros')
        else:
            messages.error(request, 'Error al actualizar el libro.')
    else:
        form = LibroForm(instance=libro)

    return render(request, 'maestros/libros/editar_libro.html', {
        'form':  form,
        'libro': libro,
    })


@admin_required
def libros_por_categoria(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    libros    = Libro.objects.filter(categoria=categoria).select_related('categoria')
    return render(request, 'maestros/libros/libros_categoria.html', {
        'categoria': categoria,
        'libros':    libros,
    })


@admin_required
def eliminar_libro(request, libro_id):
    libro = get_object_or_404(Libro, id=libro_id)
    if request.method == 'POST':
        titulo = str(libro)
        libro.delete()
        _registrar_auditoria(request, 'ELIMINAR', 'Libro', titulo)
        messages.success(request, f'Libro "{titulo}" eliminado correctamente.')
    return redirect('library:listar_libros')


@admin_required
def carga_masiva_libros(request):
    """
    GET:  Muestra el formulario vacío con instrucciones.
    POST: Procesa los archivos y muestra el resumen en la misma página.
    """
    from ..services.carga_masiva import procesar_carga_libros

    resultado = None
    form = CargaMasivaLibrosForm()

    if request.method == "POST":
        form = CargaMasivaLibrosForm(request.POST, request.FILES)
        if form.is_valid():
            from ..services.carga_masiva import procesar_actualizacion_libros
            excel_file = request.FILES['archivo_excel']
            zip_file   = request.FILES.get('archivo_zip')
            modo       = form.cleaned_data['modo']

            if modo == 'actualizar':
                resultado = procesar_actualizacion_libros(excel_file, zip_file)
                if resultado['actualizados'] > 0:
                    messages.success(request, f"{resultado['actualizados']} libro(s) actualizado(s) exitosamente.")
            else:
                resultado = procesar_carga_libros(excel_file, zip_file)
                if resultado['creados'] > 0:
                    messages.success(request, f"{resultado['creados']} libro(s) creado(s) exitosamente.")

            if resultado['errores']:
                messages.error(
                    request,
                    f"{len(resultado['errores'])} fila(s) con errores. "
                    f"Revisa el detalle a continuación."
                )
            if resultado['advertencias']:
                messages.warning(
                    request,
                    f"{len(resultado['advertencias'])} advertencia(s). "
                    f"Revisa el detalle a continuación."
                )

    return render(request, 'maestros/libros/carga_masiva_libros.html', {
        'form':      form,
        'resultado': resultado,
        'modo':      form.cleaned_data.get('modo', 'crear') if resultado else 'crear',
    })


@admin_required
def descargar_plantilla_libros(request):
    """Genera y descarga en memoria una plantilla Excel para la carga masiva de libros."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Libros"

    encabezados = ['titulo', 'autor', 'año', 'categoria', 'isbn', 'descripcion', 'archivo_pdf', 'portada']
    ws.append(encabezados)

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    ws.append([
        'El Principito',
        'Antoine de Saint-Exupéry',
        1943,
        'Literatura',
        '978-2-07-040850-4',
        'Un clásico de la literatura universal sobre la amistad y el amor.',
        'el_principito.pdf',
        'el_principito.jpg',
    ])

    anchos = [30, 30, 8, 20, 22, 50, 25, 25]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    ws2 = wb.create_sheet(title="Instrucciones")
    instrucciones = [
        ("Campo",       "¿Requerido?", "Descripción"),
        ("titulo",      "SÍ",  "Título del libro (máx. 200 caracteres)."),
        ("autor",       "SÍ",  "Nombre del autor (máx. 200 caracteres)."),
        ("año",         "SÍ",  "Año de publicación, número entre 1000 y 2100."),
        ("categoria",   "SÍ",  "Nombre exacto de la categoría (debe existir en el sistema)."),
        ("isbn",        "NO",  "ISBN del libro. Si ya existe uno igual, la fila se omite."),
        ("descripcion", "NO",  "Descripción opcional del libro."),
        ("archivo_pdf", "NO",  "Nombre del archivo PDF tal como aparece en el ZIP (con extensión)."),
        ("portada",     "NO",  "Nombre del archivo de portada tal como aparece en el ZIP (con extensión)."),
    ]
    for fila in instrucciones:
        ws2.append(fila)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 65

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_libros.xlsx"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Libros — ADMIN y LECTOR (con control de acceso por categoría)
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def leer_libro(request, autor_slug, titulo_slug):
    libro = next(
        (l for l in Libro.objects.all()
         if slugify(l.autor) == autor_slug and slugify(l.titulo) == titulo_slug),
        None,
    )
    if libro is None:
        raise Http404("Libro no encontrado.")
    if not usuario_puede_ver_libro(request.user, libro):
        messages.error(request, "No tienes acceso a este libro.")
        return redirect('library:dashboard')

    # F1 — Registrar apertura en historial
    HistorialLectura.objects.create(usuario=request.user, libro=libro)
    stats, _ = EstadisticaUsuario.objects.get_or_create(usuario=request.user)
    stats.libros_leidos = (
        HistorialLectura.objects.filter(usuario=request.user)
        .values('libro').distinct().count()
    )
    stats.save(update_fields=['libros_leidos'])

    # F5 — Cargar última página guardada
    progreso = ProgresoLectura.objects.filter(usuario=request.user, libro=libro).first()
    ultima_pagina_guardada = progreso.ultima_pagina if progreso else 1

    return render(request, 'maestros/libros/leer_libro.html', {
        'libro':                  libro,
        'ultima_pagina_guardada': ultima_pagina_guardada,
    })


@login_required
def libros_usuario(request):
    from ..models import CategoriaPermitida
    categorias_ids = CategoriaPermitida.objects.filter(
        usuario=request.user
    ).values_list('categoria', flat=True)
    libros = Libro.objects.filter(
        categoria__in=categorias_ids
    ).select_related('categoria')
    return render(request, 'libros.html', {'libros': libros})


@login_required
def libros_por_categoria_usuario(request, categoria_id):
    categoria = get_object_or_404(Categoria, id=categoria_id)
    if not usuario_puede_ver_categoria(request.user, categoria_id):
        messages.error(request, "No tienes acceso a esta categoría.")
        return redirect('library:dashboard')
    libros = Libro.objects.filter(categoria=categoria).order_by('titulo')
    return render(request, 'library/categorias/libros_por_categoria.html', {
        'categoria': categoria,
        'libros':    libros,
    })


@login_required
def detalle_libro_usuario(request, libro_id):
    from ..models import Favorito
    libro = get_object_or_404(Libro, id=libro_id)
    if not usuario_puede_ver_libro(request.user, libro):
        messages.error(request, "No tienes acceso a este libro.")
        return redirect('library:dashboard')
    es_favorito = Favorito.objects.filter(usuario=request.user, libro=libro).exists()
    return render(request, 'library/categorias/detalle_libro.html', {
        'libro':       libro,
        'es_favorito': es_favorito,
    })


@login_required
def buscar_libro(request):
    from ..models import CategoriaPermitida
    query        = request.GET.get('q', '').strip()
    categoria_id = request.GET.get('categoria', '').strip()
    year_desde   = request.GET.get('year_desde', '').strip()
    year_hasta   = request.GET.get('year_hasta', '').strip()
    orden        = request.GET.get('orden', 'az').strip()

    orden_campo = _ORDER_MAP.get(orden, 'titulo')
    libros = Libro.objects.select_related('categoria')

    if not es_admin(request.user):
        categorias_ids = CategoriaPermitida.objects.filter(
            usuario=request.user
        ).values_list('categoria_id', flat=True)
        libros = libros.filter(categoria_id__in=categorias_ids)
        categorias_disponibles = Categoria.objects.filter(id__in=categorias_ids)
    else:
        categorias_disponibles = Categoria.objects.all()

    if query:
        libros = libros.filter(Q(titulo__icontains=query) | Q(autor__icontains=query))
    if categoria_id:
        libros = libros.filter(categoria_id=categoria_id)
    if year_desde:
        try:
            libros = libros.filter(año__gte=int(year_desde))
        except ValueError:
            pass
    if year_hasta:
        try:
            libros = libros.filter(año__lte=int(year_hasta))
        except ValueError:
            pass

    libros = libros.order_by(orden_campo)

    return render(request, 'library/buscar_libro.html', {
        'libros':                 libros,
        'query':                  query,
        'resultados_count':       libros.count(),
        'categorias_disponibles': categorias_disponibles,
        'filtros': {
            'categoria_id': categoria_id,
            'year_desde':   year_desde,
            'year_hasta':   year_hasta,
            'orden':        orden,
        },
    })
